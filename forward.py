import string
import time
import random
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, load_workbook

base_url = 'https://www.editus.lu/fr/particulier/lettre-'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.67 Safari/537.36'
}

wb = load_workbook('juice.xlsx')
ws = wb.active

page_urls = ws['A']

for i, page in enumerate(page_urls):
    if i < 10:
        try:
            print(page.value)
            response = requests.get(page.value)

            if response.status_code == 200:
                soup = bs(response.content, 'html.parser')
                main_info = soup.find('section', id = 'main-info')
                
                person_name = main_info.find('h1', class_ = 'name').text.strip()
                ws['B{}'.format(i + 1)].value = person_name

                address = main_info.find('span', class_ = 'address')
                street = address.find('span', class_ = 'street').text.strip()
                ws['C{}'.format(i + 1)].value = street
                zip_code = address.find('span', class_ = 'zip-code').text.strip()
                ws['D{}'.format(i + 1)].value = zip_code
                locality = address.find('span', class_ = 'locality').text.strip()
                ws['E{}'.format(i + 1)].value = locality

                occupation = main_info.find('ul', class_ = 'additionnal-info').find('h2', class_ = 'label').text.strip()
                ws['F{}'.format(i + 1)].value = occupation

                phone_number = main_info.find('div', class_ = 'buttons').find('span', class_ = 'phone-number').text.strip()
                ws['G{}'.format(i + 1)].value = occupation
        except Exception as error:
            print(error)
            pass
            
        time.sleep(random.randint(1, 3))

    else:
        break

try:
    wb.save('output13.xlsx')
except Exception as error:
    print(error)
    pass